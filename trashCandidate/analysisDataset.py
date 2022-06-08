# 分析原始数据，生成excel报表，展现每个视频和标注的情况。如视频时长、FPS，各类别标注量等
import numpy as np
import os
import cv2
from tqdm import tqdm
from openpyxl import Workbook
from time import strftime
from time import gmtime

vid_folder = 'E:/Research/2021TrafficSceneClassification/Datasets/HSD/FINALDATA/VIDS'
csv_folder = 'E:/Research/2021TrafficSceneClassification/Datasets/HSD/FINALDATA/CSV'

cname = ['Highway','Local','Ramp','Urban','Rural','Unknow']


train_video = 'E:/Research/2021TrafficSceneClassification/Datasets/HSD/FINALDATA/trainsplit.txt'
val_video = 'E:/Research/2021TrafficSceneClassification/Datasets/HSD/FINALDATA/valsplit.txt'
with open(train_video) as f:
    train_video = f.readlines()
    for i in range(len(train_video)):
        if train_video[i][-1] == '\n':
            train_video[i] = train_video[i][:-1] 
    f.close()
with open(val_video) as f:
    val_video = f.readlines()
    for i in range(len(val_video)):
        if val_video[i][-1] == '\n':
            val_video[i] = val_video[i][:-1] 
    f.close()


videos = []
for _,_,f in os.walk(vid_folder):
    for i in f:
        if i.endswith('mp4'):
            videos.append(i[:-4])
pass

wb = Workbook()
ws = wb.active
ws['A1'] = 'Video'
ws['B1'] = 'Highway'
ws['C1'] = 'Local'
ws['D1'] = 'Ramp'
ws['E1'] = 'Urban'
ws['F1'] = 'Rural'
ws['G1'] = 'Unknown'
ws['H1'] = 'Split'
ws['I1'] = 'FPS'
ws['J1'] = 'Fno'
ws['K1'] = 'Time'

pbar = tqdm(enumerate(videos))
for vidx,name in pbar:

    csv = os.path.join(csv_folder,name+'.csv')
    with open(csv) as f:
        lines = f.readlines()
        f.close()
        lines = lines[1:]
    cap = os.path.join(vid_folder,name+'.mp4')
    cap = cv2.VideoCapture(cap)
    fno = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    fps = int(cap.get(cv2.CAP_PROP_FPS))
    time = int(fno/fps)
    time = strftime("%H:%M:%S", gmtime(time))
    cap.release()
    assert(fno == len(lines))

    mydict = {}

    for i,line in enumerate(lines):
        l = int(line.split(',')[-2])       
        if l == 0:
            res = -1 # unknown
        elif l == 1:
            res = 1 # local
        elif l == 2:
            res = 0 # highway
        elif l == 3:
            res = 2 # ramp
        elif l == 4:
            res = 3 # urban
        elif l == 5:
            res = 4 # rural
        else:
            print(l)
            print('Error! No this label!')
        if res not in mydict.keys():
            mydict[res] = 0
        mydict[res] += 1

    mydict = sorted(mydict.items(), key = lambda kv:(kv[0]))

    ws['A'+str(vidx+2)] = name

    row = str(vidx+2)

    for m in mydict:
        c = m[0]
        v = m[1]
        if c == -1:
            c = 5
        pos = chr(c+1+ord('A')) + str(vidx+2)
        ws[pos] = v
    
    name += '.mp4'
    if name in train_video:
        ws['H'+str(vidx+2)] = 'train'
    if name in val_video:
        ws['H'+str(vidx+2)] = 'val'
    ws['I'+row] = fps
    ws['j'+row] = fno
    ws['k'+row] = time
    

wb.save('E:/Research/2021TrafficSceneClassification/Datasets/HSD/datasetAnalysis.xlsx')
wb.close()