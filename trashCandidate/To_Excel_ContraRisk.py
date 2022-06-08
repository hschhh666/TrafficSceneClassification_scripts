import numpy as np
import os
import openpyxl
from openpyxl import Workbook
from tqdm import tqdm

data_path = 'E:/Research/2021TrafficSceneClassification/Datasets/HSD/experimentRes/Contras/train'

f = []
for _,_,tmpf in os.walk(data_path):
    for i in tmpf:
        if i[-3:] == 'mp4':
            f.append(i[:-4])

for video_name in f:
    risk_name = video_name+'_to_classes_risk'
    if os.path.exists(os.path.join(data_path, risk_name + '.npy')) and not os.path.exists(os.path.join(data_path, video_name + '_to_classes_risk.xlsx')):



        risk = np.load(os.path.join(data_path, risk_name + '.npy'))


        wb = Workbook()

        ws = wb.active

        ws['A1'] = 'Highway(Red)'
        ws['B1'] = 'Local(Green)'
        ws['C1'] = 'Ramp(Purple)'
        ws['D1'] = 'Urban(Blue)'
        char = ['A','B','C','D']

        datanum = np.shape(risk)[0]

        pbar = tqdm(range(datanum), desc=video_name)
        for i in pbar:
            for j in range(4):
                pos = char[j] + str(i+2)
                value = risk[i][j]
                ws[pos] = value



        wb.save(os.path.join(data_path, video_name + '_to_classes_risk.xlsx'))
        wb.close()

print('Done')